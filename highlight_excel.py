import openpyxl
import sys
from openpyxl.styles import PatternFill

def highlight_candidate(candidate_name, excel_path):
    try:
        # Load the workbook and sheet
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active

        # Define highlight and normal fill
        highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        normal_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # Reset all rows to normal state
        for row in sheet.iter_rows(min_row=2):  # Assuming headers are in the first row
            for cell in row:
                cell.fill = normal_fill

        # Find and highlight the row matching the candidate name
        row_found = False
        for row in sheet.iter_rows(min_row=2):  # Assuming headers are in the first row
            if row[1].value == candidate_name:  # Assuming 'Name' is in the second column
                row_found = True
                for cell in row:
                    cell.fill = highlight_fill
                break

        if not row_found:
            print(f"Candidate '{candidate_name}' not found in the Excel sheet.")
        else:
            # Save and open the workbook
            wb.save(excel_path)
            print(f"Highlighted row for Candidate: {candidate_name}")
            import os
            os.system(f'start excel "{excel_path}"')

    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    candidate_name = sys.argv[1]  # Candidate name passed from the Node.js server
    excel_path = sys.argv[2]     # Path to the Excel file
    highlight_candidate(candidate_name, excel_path)
